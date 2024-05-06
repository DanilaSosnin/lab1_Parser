from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

def parse_chitaigorod():
    fn = 'ParserOutput.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    index = 2
    ws['A1'] = 'Название'
    ws['B1'] = 'Цена'
    ws['C1'] = 'Автор'

    for numpage in range(1, 7):
        url = ('https://www.chitai-gorod.ru/search?phrase=python&page=' + str(numpage)) # передаем необходимы URL адрес
        page = requests.get(url) # отправляем запрос методом Get на данный адрес и получаем ответ в переменную
        print("Просканирована страница", numpage, "из 6, статус", page.status_code) # смотрим ответ
        soup = BeautifulSoup(page.text, "html.parser") # передаем страницу в bs4

        block = soup.findAll('article', class_='product-card product-card product')

        for data in block:
            if data.find('div', {'class':'product-price__value product-price__value--discount'}):
                price = data.find('div', {'class':'product-price__value product-price__value--discount'}).text.replace("&nbsp;", '').replace('\n', '').replace(' ', '')
            elif data.find('div', {'class':'product-price__value'}):
                price = data.find('div', {'class': 'product-price__value'}).text.replace("&nbsp;", '').replace('\n', '').replace(' ', '')
            else:
                price = "Цена не указана"
            if data.find('div', {'class':'product-title__head'}):
                name = data.find('div', {'class':'product-title__head'}).text.replace("\n", '')
            else:
                name = "Без названия"
            if data.find('div', {'class':'product-title__author'}):
                author = data.find('div', {'class':'product-title__author'}).text.replace("\n", '')
                if author == '':
                    author = 'Автор не указан'
            else:
                author = "None"
            ws['A' + str(index)] = name
            ws['B' + str(index)] = price
            ws['C' + str(index)] = author
            index += 1
            print(price, name, author, sep=" ")

    wb.save(fn)
    wb.close()